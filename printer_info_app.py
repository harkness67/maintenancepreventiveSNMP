# printer_info_app.py

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import pandas as pd
from datetime import datetime
from snmp_helper import get_printer_info
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import re
import subprocess
import webbrowser
import smtplib
import socket
import numpy as np
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email_helper import send_email_alert, send_email_alert_serial_number_change
import logging
from tkinter import simpledialog
from tkinter import font
from datetime import datetime
from openpyxl.utils import get_column_letter


logging.basicConfig(filename='app_log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class PrinterInfoApp:
    def __init__(self, master):
        self.master = master
        self.master.title('Outil de collecte SNMP pour imprimantes sensibles aux HUS')
        self.master.geometry('1450x800')
        
        # Création des onglets
        self.tab_control = ttk.Notebook(master)
        
        # Onglet de collecte d'informations
        self.tab_collect = ttk.Frame(self.tab_control)
        self.tab_settings = ttk.Frame(self.tab_control)
        self.tab_control.add(self.tab_collect, text='Collecte d\'informations')
        self.tab_control.add(self.tab_settings, text='Paramètres')
        self.tab_control.bind("<<NotebookTabChanged>>", self.on_tab_changed)

        
        
        # Cadre pour la gestion des adresses IP
        self.ip_management_frame = ttk.LabelFrame(self.tab_settings, text="Gestion des adresses IP et seuils de déclenchement")
        self.ip_management_frame.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)

        # Onglet de paramètres
        self.ip_canvas = tk.Canvas(self.ip_management_frame)  
        self.ip_canvas.grid(row=1, column=0, sticky='nsew', padx=10, pady=10)

        self.ip_scrollbar = ttk.Scrollbar(self.ip_management_frame, orient="vertical", command=self.ip_canvas.yview)
        self.ip_scrollbar.grid(row=1, column=1, sticky='ns')

        self.ip_canvas.config(yscrollcommand=self.ip_scrollbar.set)

        self.ip_treeview = ttk.Treeview(self.ip_canvas, columns=('IP Address', 'Threshold'), height=20)  # Fixe la hauteur du Treeview à 20 lignes
        self.ip_treeview_window = self.ip_canvas.create_window((0, 0), window=self.ip_treeview, anchor="nw")

        self.ip_treeview.bind("<Configure>", lambda event: self.ip_canvas.configure(scrollregion=self.ip_canvas.bbox("all")))
        self.ip_treeview.bind("<Double-1>", self.on_item_double_click)


        self.ip_treeview.column('IP Address', width=100)  # Ajustez la largeur selon vos besoins
        self.ip_treeview.column('Threshold', width=100)  # Ajustez la largeur selon vos besoins

        self.ip_treeview['show'] = 'headings'
        self.ip_treeview.heading('IP Address', text='Adresse IP')
        self.ip_treeview.heading('Threshold', text='Seuil paramétré')

        self.ip_management_frame.grid_columnconfigure(0, weight=1)

        # Récupérez la hauteur du Treeview et ajustez la hauteur du Canvas en conséquence
        tree_height = self.ip_treeview.winfo_height()
        self.ip_canvas.configure(height=tree_height)


      
        
        # Champ de saisie pour la nouvelle adresse IP
        self.new_ip_label = ttk.Label(self.ip_management_frame, text="Nouvelle adresse IP :")
        self.new_ip_label.grid(row=2, column=0, sticky='w', padx=10)
        self.new_ip_entry = ttk.Entry(self.ip_management_frame)
        self.new_ip_entry.grid(row=2, column=1, sticky='w', padx=10)

        # Champ de saisie pour le nouveau seuil
        self.new_threshold_label = ttk.Label(self.ip_management_frame, text="Nouveau seuil :")
        self.new_threshold_label.grid(row=3, column=0, sticky='w', padx=10)
        self.new_threshold_entry = ttk.Entry(self.ip_management_frame)
        self.new_threshold_entry.grid(row=3, column=1, sticky='w', padx=10)


        # Bouton pour charger le contenu du fichier
        #self.load_button = ttk.Button(self.ip_management_frame, text="Charger manuellement", command=self.load_ips)
        #self.load_button.grid(row=0, column=0, pady=10, padx=10, sticky='w', columnspan=2)


        # Bouton pour sauvegarder les modifications dans le fichier
        self.save_button = ttk.Button(self.ip_management_frame, text="Sauvegarder", command=self.save_ips)
        self.save_button.grid(row=3, column=2, pady=10, padx=10, sticky='e')

        
        # Bouton pour ajouter une adresse IP et un seuil
        self.add_button = ttk.Button(self.ip_management_frame, text="Ajouter", command=self.add_ip)
        self.add_button.grid(row=1, column=2, pady=10, padx=10, sticky='ns')

        # Bouton pour supprimer une adresse IP sélectionnée et son seuil
        self.remove_button = ttk.Button(self.ip_management_frame, text="Supprimer", command=self.remove_ip)
        self.remove_button.grid(row=1, column=3, pady=10, padx=10, sticky='ns')
    
        
        # Onglet de rapport
        self.tab_report = ttk.Frame(self.tab_control)
        self.tab_control.add(self.tab_report, text='Rapport')
        self.log_text_widget = tk.Text(self.tab_report, height=20, width=80)
        self.log_text_widget.pack(padx=10, pady=10, fill='both', expand=True)
        
        self.refresh_log_button = ttk.Button(self.tab_report, text='Actualiser le log', command=self.load_log_content)
        self.refresh_log_button.pack(pady=10)
        
        self.clear_log_button = ttk.Button(self.tab_report, text='Effacer le log', command=self.clear_log)
        self.clear_log_button.pack(pady=10)

        
        self.tab_control.pack(expand=1, fill='both')
        
        # Cadre d'informations
        self.info_frame = ttk.LabelFrame(self.tab_collect, text="Informations")
        self.info_frame.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
        
        # Création des widgets pour le cadre d'informations
        self.label = ttk.Label(self.info_frame, text="Le fichier contenant les adresses ip se situe à la racine du dossier et se nomme 'adresses_ip.txt'")
        self.label.grid(row=0, column=0, sticky='w', padx=10, pady=5)

        self.file_text_label = ttk.Label(self.info_frame, text='')
        self.file_text_label.grid(row=1, column=0, sticky='w', padx=10, pady=5)

        self.button = ttk.Button(self.info_frame, text='Obtenir les informations', command=self.get_info)
        self.button.grid(row=2, column=0, pady=10, padx=10, sticky='w')

        # Cadre d'actions
        self.action_frame = ttk.LabelFrame(self.tab_collect, text="Actions")
        self.action_frame.grid(row=1, column=0, sticky='nsew', padx=10, pady=10)

        self.clear_ui_button = ttk.Button(self.action_frame, text='Effacer les relevés (Interface)', command=self.clear_ui_data)
        self.clear_ui_button.grid(row=0, column=0, pady=10, padx=10, sticky='w')

        self.clear_excel_button = ttk.Button(self.action_frame, text='Effacer les relevés (Excel)', command=self.clear_excel_data)
        self.clear_excel_button.grid(row=1, column=0, pady=10, padx=10, sticky='w')

        # Cadre d'affichage
        self.display_frame = ttk.LabelFrame(self.tab_collect, text="Affichage")
        self.display_frame.grid(row=2, column=0, sticky='nsew', padx=10, pady=10)

        self.table = ttk.Treeview(self.display_frame, columns=('Name', 'Address', 'Serial Number', 'Page Counter', 'Collect Date', 'First Collect Date', 'diff'))
        self.table = ttk.Treeview(self.display_frame, columns=('Name', 'Address', 'Serial Number', 'Page Counter', 'Collect Date', 'First Collect Date', 'diff', 'Percentage'))

        self.table.heading('Name', text='Nom réseau')
        self.table.heading('Address', text='Adresse IP')
        self.table.heading('Serial Number', text='Numéro de série')
        self.table.heading('Page Counter', text='Compteur de pages')
        self.table.heading('Collect Date', text='Date de collecte')
        self.table.heading('First Collect Date', text='Date de la première collecte')
        self.table.heading('diff', text='Écart') 
        self.table.heading('Percentage', text='% d\'atteinte du seuil')
        self.table.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
        
        self.table['show'] = 'headings'

        
        for col in self.table['columns']:
            self.table.column(col, anchor='center')
            self.table.column("Name", width=150)
            self.table.column("Address", width=150)
            self.table.column("Serial Number", width=150)
            self.table.column("Page Counter", width=150)
            self.table.column("Collect Date", width=150)
            self.table.column("First Collect Date", width=200)
            self.table.column("diff", width=150)
            self.table.column("Percentage", width=150)

            values = [self.table.set(child, col) for child in self.table.get_children()]
            if values:  # Vérifiez si la liste des valeurs n'est pas vide
                max_width = max(self.table.bbox(item, col)[2] for item in self.table.get_children())
                self.table.column(col, width=max_width)

        
        # Configurer le redimensionnement automatique des colonnes et des lignes
        self.tab_collect.grid_columnconfigure(0, weight=1)
        self.tab_collect.grid_rowconfigure(0, weight=1)
        self.tab_collect.grid_rowconfigure(1, weight=1)
        self.tab_collect.grid_rowconfigure(2, weight=1)

        self.open_excel_button = ttk.Button(self.tab_collect, text='Ouvrir le fichier Excel', command=self.open_excel)
        self.open_excel_button.grid(row=3, column=0, pady=10, padx=10, sticky='w')
        
        self.alert_thresholds = {}
        # Dictionnaire pour stocker la dernière valeur du compteur de pages par adresse IP
        self.last_counter = {}
        self.report = {"success": [], "errors": []}  # Initialisation du rapport

    def update_scrollregion(self, event):
        self.ip_canvas.configure(scrollregion=self.ip_canvas.bbox("all"))
        
    def scroll_treeview(self, *args):
        self.ip_treeview.yview(*args)

        self.ip_canvas.bind('<MouseWheel>', self.scroll_treeview)
        self.ip_canvas.bind("<Configure>", resize_treeview)

        
    def resize_treeview(self, event):
        # Si la largeur du canvas est plus grande que la largeur minimale du Treeview, ajustez la largeur du Treeview
        if event.width > self.ip_treeview.winfo_reqwidth():
            self.ip_treeview.config(width=event.width)


    def on_tab_changed(self, event):
        selected_tab = self.tab_control.select()
        tab_text = self.tab_control.tab(selected_tab, "text")
    
        if tab_text == "Paramètres":
            self.load_ips()
        elif tab_text == "Rapport":
            self.load_log_content()


    def load_log_content(self):
        try:
            with open("printer_retrieval.log", "r") as f:
                content = f.read()
            self.log_text_widget.delete(1.0, tk.END)  # Effacer tout le contenu existant
            self.log_text_widget.insert(tk.END, content)  # Insérer le contenu du fichier
        except FileNotFoundError:
            tk.messagebox.showerror("Erreur", "Le fichier printer_retrieval.log n'a pas été trouvé.")
        except Exception as e:
            tk.messagebox.showerror("Erreur", f"Une erreur s'est produite lors de la lecture du fichier : {e}")


    def load_ips(self):
        try:
            self.ip_treeview.delete(*self.ip_treeview.get_children())  # Effacer les entrées existantes
            with open("adresses_ip.txt", "r") as f:
                lines = f.readlines()
                for line in lines:
                    ip, threshold = line.strip().split(',')
                    self.ip_treeview.insert('', 'end', values=(ip, threshold))
        except FileNotFoundError:
            tk.messagebox.showerror("Erreur", "Le fichier adresses_ip.txt n'a pas été trouvé.")
        except Exception as e:
            tk.messagebox.showerror("Erreur", f"Une erreur s'est produite lors de la lecture du fichier : {e}")

    def save_ips(self):
        try:
            content = []
            for child in self.ip_treeview.get_children():
                item = self.ip_treeview.item(child, 'values')
                content.append(','.join(item))
            with open("adresses_ip.txt", "w") as f:
                f.write('\n'.join(content))
            tk.messagebox.showinfo("Succès", "Les modifications ont été sauvegardées avec succès.")
        except Exception as e:
            tk.messagebox.showerror("Erreur", f"Une erreur s'est produite lors de la sauvegarde du fichier : {e}")

    def on_item_double_click(self, event):
        item = self.ip_treeview.selection()[0]  # Obtenez l'élément sélectionné
        col = self.ip_treeview.identify_column(event.x)  # Identifiez la colonne

        # Assurez-vous qu'un élément est bien sélectionné
        if item:
            if col == "#1":  # Si l'utilisateur a double-cliqué sur la première colonne (IP Address)
                # Vous pouvez afficher une boîte de dialogue pour obtenir la nouvelle valeur
                new_value = simpledialog.askstring("Modifier", "Entrez la nouvelle adresse IP :")
                if new_value:  # Si l'utilisateur entre une nouvelle valeur
                    confirm = tk.messagebox.askyesno("Confirmation", "Voulez-vous sauvegarder cette modification ?")
                    if confirm:
                        self.ip_treeview.set(item, 'IP Address', new_value)  # Mettez à jour la valeur dans le Treeview
                        self.save_ips()  # Sauvegardez les modifications

            elif col == "#2":  # Si l'utilisateur a double-cliqué sur la deuxième colonne (Threshold)
                new_value = simpledialog.askstring("Modifier", "Entrez le nouveau seuil :")
                if new_value:
                    confirm = tk.messagebox.askyesno("Confirmation", "Voulez-vous sauvegarder cette modification ?")
                    if confirm:
                        self.ip_treeview.set(item, 'Threshold', new_value)
                        self.save_ips()  # Sauvegardez les modifications

                
                
    def clear_log(self):
        # Effacer le contenu du fichier log
        with open("printer_retrieval.log", "w") as f:
            f.write("")
        # Effacer le contenu affiché
        self.log_text_widget.delete(1.0, tk.END)
        messagebox.showinfo("Succès", "Le log a été effacé avec succès.")
        
    def add_ip(self):
        new_ip = self.new_ip_entry.get().strip()
        try:
            new_threshold = int(self.new_threshold_entry.get().strip())
        except ValueError:
            messagebox.showerror("Erreur", "Seuil invalide. Veuillez entrer un nombre.")
            return

        if new_ip:
            # Ajout à l'interface visuelle
            self.ip_treeview.insert('', 'end', values=(new_ip, new_threshold))
            
            # Demande de confirmation pour sauvegarder
            save_changes = messagebox.askyesno("Confirmation", "Voulez-vous sauvegarder les modifications ?")
            
            # Consigner l'ajout dans le log
            logging.info(f"Ajout de l'adresse IP : {new_ip} avec un seuil de : {new_threshold}")
            
            if save_changes:
                self.save_ips()
            
            # Effacez les champs d'entrée pour une saisie future
            self.new_ip_entry.delete(0, tk.END)
            self.new_threshold_entry.delete(0, tk.END)

                
    def remove_ip(self):
        selected_item = self.ip_treeview.selection()[0]  # Obtenez l'élément sélectionné
        ip_to_delete, threshold_to_delete = self.ip_treeview.item(selected_item, "values")
        self.ip_treeview.delete(selected_item)
        # Consigner la suppression dans le log
        logging.info(f"Suppression de l'adresse IP : {ip_to_delete} avec un seuil de : {threshold_to_delete}")
        save_changes = messagebox.askyesno("Confirmation", "Voulez-vous sauvegarder les modifications ?")
        if save_changes:
            self.save_ips()




    def remove_unwanted_sheets(self):
        # Supprimer les feuilles indésirables "Sheet" et "Informations Imprimantes"
        unwanted_patterns = ['Sheet', 'Informations Imprimantes']
        
        # Charger le classeur Excel existant
        try:
            self.wb = load_workbook('printer_info.xlsx')
        except FileNotFoundError:
            # Si le fichier n'existe pas, créez-en un nouveau
            self.wb = Workbook()
        
        # Créer une liste des feuilles à conserver
        sheets_to_keep = [sheet_name for sheet_name in self.wb.sheetnames if not any(pattern in sheet_name for pattern in unwanted_patterns)]
        
        # Supprimer toutes les feuilles existantes du classeur
        for sheet_name in self.wb.sheetnames:
            self.wb.remove(self.wb[sheet_name])
        
        # Créer les feuilles à conserver dans le classeur
        for sheet_name in sheets_to_keep:
            self.wb.create_sheet(sheet_name)            
                    
    def remove_printer_sheet(self, ip_address):
        if ip_address in self.printers_data['Address'].values:
            index = self.printers_data[self.printers_data['Address'] == ip_address].index[0]
            self.printers_data.drop(index, inplace=True)


    def clear_ui_data(self):
        # Effacer les données dans le widget de texte
        self.log_text_widget.delete('1.0', tk.END)


        # Effacer les données dans le tableau
        self.table.delete(*self.table.get_children())

        # Effacer les valeurs du dernier relevé
        self.last_counter = {}


    def clear_excel_data(self):
        # Effacer les données dans le DataFrame printers_data
        self.printers_data = pd.DataFrame(columns=('Nom réseau', 'Adresse IP', 'Numéro de série', 'Compteur de pages', 'Date de collecte', 'Écart', 'Première relève', 'Dernière relève', 'Date de la première relève', 'Nombre de relevés', 'Commentaire'))


        # Enregistrer les données dans le fichier Excel (ceci va écraser le contenu du fichier)
        self.printers_data.to_excel('printer_info.xlsx', index=False)

                
    def open_excel(self):
        try:
            file_path = os.path.abspath('printer_info.xlsx')
            webbrowser.open(file_path)
        except Exception as e:
            messagebox.showerror('Erreur', f'Impossible d\'ouvrir le fichier Excel : {str(e)}')
    

    def get_info(self):
        # Récupérer le chemin du fichier contenant les adresses IP cibles
        file_path = 'adresses_ip.txt'
        
        # Initialisation par défaut des valeurs utilisées dans la fonction
        percentage = 0 
        
        # Dictionnaire pour stocker les valeurs actuelles de "Première relève" et "Date de la première relève"
        first_readings = {}
        date_first_readings = {}
        
        try:
            with open(file_path, 'r') as file:
                lines = [line.strip().split(',') for line in file.readlines()]
                ip_addresses = [line[0] for line in lines]
                thresholds = [int(line[1]) for line in lines]
                self.alert_thresholds = dict(zip(ip_addresses, thresholds))
        except FileNotFoundError:
            messagebox.showerror('Erreur', f'Le fichier {file_path} n\'existe pas.')
            return
        
        # Charger les données existantes à partir du fichier Excel s'il existe
        try:
            self.printers_data = pd.read_excel('printer_info.xlsx')
        except FileNotFoundError:
            # Si le fichier n'existe pas, initialiser avec un DataFrame vide
            columns = ['Nom réseau', 'Adresse IP', 'Numéro de série', 'Compteur de pages', 
                       'Date de collecte', 'Écart', 'Première relève', 'Dernière relève', 'Date de la première relève', 'Nombre de relevés', 'Commentaire', 'Pourcentage Atteinte']
            self.printers_data = pd.DataFrame(columns=columns)
            
        if 'Nombre de relevés' not in self.printers_data.columns:
            self.printers_data['Nombre de relevés'] = 1

        
        for ip in ip_addresses:
            # Initialisation par défaut
            date_first_readings[ip] = "N/A"
            
            if ip in self.printers_data['Adresse IP'].values:
                index = self.printers_data[self.printers_data['Adresse IP'] == ip].index[0]
                first_readings[ip] = self.printers_data.at[index, 'Première relève']
                date_first_readings[ip] = self.printers_data.at[index, 'Date de la première relève']

        
        # Créer une liste pour stocker les informations de chaque imprimante
        printer_info_list = []
        
        for ip_address in ip_addresses:
            new_row = None  # Initialisation de new_row pour éviter l'erreur UnboundLocalError
            date_first_reading = date_first_readings.get(ip_address, "N/A")


            # Récupérer les informations de l'imprimante
            try:
                info = get_printer_info(ip_address)
            except Exception as e:
                logging.error(f"Erreur lors de la collecte d'informations pour l'adresse IP {ip_address}. Erreur : {e}")
                continue
            
            if info is None:
                messagebox.showerror('Erreur', f'Impossible d\'obtenir les informations de l\'imprimante à l\'adresse {ip_address}')
                error_msg = f"Erreur lors de la récupération des informations de l'imprimante à l'adresse {ip_address}"
                self.report["errors"].append(error_msg)
                logging.error(error_msg)
                continue
            
            success_msg = f"Récupération des informations pour l'imprimante à l'adresse {ip_address} réussie."
            self.report["success"].append(success_msg)
            logging.info(success_msg)
            
            name, serial_number, page_counter, collect_date = info
            page_counter = int(page_counter)
            
            # Vérifier si les données de relevé existent déjà dans le DataFrame
            existing_index = self.printers_data[self.printers_data['Adresse IP'] == ip_address].index
            
            if not existing_index.empty:
                index = existing_index[0]
                
                # Si le numéro de série a changé, ajoutez une nouvelle ligne
                if self.printers_data.at[index, 'Numéro de série'] != serial_number:
                    last_serial = self.printers_data.at[index, 'Numéro de série']
                    self.handle_serial_number_change(ip_address, name, last_serial, serial_number)
                    new_row = {
                        'Nom réseau': name, 
                        'Adresse IP': ip_address, 
                        'Numéro de série': serial_number, 
                        'Compteur de pages': page_counter, 
                        'Date de collecte': collect_date, 
                        'Écart': 0,
                        'Première relève': page_counter, 
                        'Dernière relève': page_counter,
                        'Date de la première relève': collect_date,
                        'Nombre de relevés': 1,
                        'Commentaire': 'Numéro de série différent détecté, nouvelle ligne créée pour cette FA', 
                        'Pourcentage Atteinte': f"{percentage:.2f}%"  # Assurez-vous de formater le pourcentage correctement.
                    }
                    self.printers_data = self.printers_data.append(new_row, ignore_index=True)
                    messagebox.showinfo('Information', f'Le numéro de série de l\'imprimante {name} à l\'adresse {ip_address} a changé. Une nouvelle entrée a été ajoutée.')
                    logging.info(f"Le numéro de série de l'imprimante {name} à l'adresse {ip_address} a changé. Une nouvelle entrée a été ajoutée.")
                else:
                    # Mettre à jour les données existantes
                    self.printers_data.at[index, 'Compteur de pages'] = page_counter
                    self.printers_data.at[index, 'Date de collecte'] = collect_date
                    self.printers_data.at[index, 'Dernière relève'] = page_counter


                    
                    if pd.isna(self.printers_data.at[index, 'Première relève']):
                        self.printers_data.at[index, 'Première relève'] = page_counter
                        self.printers_data.at[index, 'Date de la première relève'] = collect_date
                    
                    # Calculer l'écart
                    first_reading = self.printers_data.at[index, 'Première relève']
                    diff = page_counter - first_reading
                    threshold = self.alert_thresholds.get(ip_address, 0)  # Récupérer le seuil pour l'adresse IP actuelle

                    # Calculer le pourcentage d'atteinte du seuil
                    if threshold != 0:  # Vérifier que le seuil n'est pas zéro pour éviter la division par zéro
                        percentage = (diff / threshold) * 100 if threshold != 0 else 0
                    else:
                        percentage = 0  # Assigner 0 si le seuil est zéro
                        
                    self.printers_data.at[index, 'Pourcentage Atteinte'] = f"{percentage:.2f}%"

                    

                                                          
                    # Vérifier si l'écart atteint le seuil d'alerte et déclencher l'alerte si nécessaire
                    self.handle_alert(ip_address, name, diff)
                    
                    # Incrémentation du "Nombre de relevés"
                    self.printers_data.at[index, 'Nombre de relevés'] = self.printers_data.at[index, 'Nombre de relevés'] + 1
                


                # Insérer les informations dans le tableau
                self.table.insert('', 'end', values=(name, ip_address, serial_number, page_counter, collect_date, date_first_readings[ip_address], diff, f"{percentage:.2f}%"))
            
            else:
                # Ajouter une nouvelle entrée pour l'imprimante
                new_row = {
                    'Nom réseau': name, 
                    'Adresse IP': ip_address, 
                    'Numéro de série': serial_number, 
                    'Compteur de pages': page_counter, 
                    'Date de collecte': collect_date, 
                    'Écart': 0,  # Écart initialisé à 0 pour une nouvelle entrée
                    'Première relève': page_counter, 
                    'Dernière relève': page_counter,
                    'Date de la première relève': collect_date,
                    'Nombre de relevés': 1,
                    'Commentaire': '',
                    'Pourcentage Atteinte': f"{percentage:.2f}%",

                }
                printer_info_list.append(new_row)

        
        # Mettre à jour le DataFrame avec les nouvelles données
        if printer_info_list:
            new_data = pd.DataFrame(printer_info_list)
            self.printers_data = pd.concat([self.printers_data, new_data], ignore_index=True)
        
        # Supprimer les doublons
        self.printers_data.drop_duplicates(subset=['Adresse IP'], keep='first', inplace=True)
        
        # Réaffectez les valeurs de "Première relève" et "Date de la première relève" à partir du dictionnaire
        for ip, value in first_readings.items():
            if ip in self.printers_data['Adresse IP'].values:
                index = self.printers_data[self.printers_data['Adresse IP'] == ip].index[0]
                self.printers_data.at[index, 'Première relève'] = value
                self.printers_data.at[index, 'Date de la première relève'] = date_first_readings[ip]
       
       
        # Juste avant d'enregistrer le DataFrame dans le fichier Excel :
        ordered_columns = ['Nom réseau', 'Adresse IP', 'Numéro de série', 'Compteur de pages', 
                           'Date de collecte', 'Écart', 'Première relève', 'Dernière relève', 
                           'Date de la première relève', 'Nombre de relevés', 'Pourcentage Atteinte']

        self.printers_data = self.printers_data[ordered_columns]

        
        
        
        
        
      
        # Enregistrer les données dans le fichier Excel
        self.printers_data.to_excel('printer_info.xlsx', index=False)

        #message qui informe du succès de la relève
        tk.messagebox.showinfo('Succès', 'La relève a été réalisée avec succès.')
        
        # Appel de finalize_report pour générer le rapport
        self.finalize_report()

        
    def handle_alert(self, ip_address, printer_name, diff):
        # Vérifier si l'écart atteint le seuil d'alerte
        if ip_address in self.alert_thresholds:
            threshold = self.alert_thresholds[ip_address]
            if diff >= threshold:
                # Appeler la méthode send_email_alert pour envoyer l'e-mail d'alerte
                email_sent = send_email_alert(ip_address, printer_name, diff, self.alert_thresholds)  # Ajoutez cette ligne
                if email_sent:
                    messagebox.showinfo('Information', f'Un e-mail d\'alerte a été envoyé pour l\'imprimante {printer_name} à l\'adresse {ip_address}.')

                # Afficher une alerte dans l'interface utilisateur
                messagebox.showwarning('Alerte', f"Le compteur de pages de l'imprimante {printer_name} à l'adresse {ip_address} a atteint un écart de {diff}.")
                logging.warning(f"Alerte: Le compteur de pages de l'imprimante {printer_name} à l'adresse {ip_address} a atteint un écart de {diff}.")


                # Mettre en évidence l'imprimante dans le tableau
                item_id = self.get_item_id(ip_address)
                if item_id:
                    self.table.item(item_id, tags=('alert',))


    def get_item_id(self, ip_address):
        for item in self.table.get_children():
            if self.table.item(item, "values")[1] == ip_address:  # Assumer que l'adresse IP est la 2ème colonne
                return item
        return None


    def handle_serial_number_change(self, ip_address, printer_name, last_serial_number, current_serial_number):
        send_email_alert_serial_number_change(ip_address, printer_name, last_serial_number, current_serial_number)
        messagebox.showwarning('Alerte', f"Le numéro de série de l'imprimante {printer_name} à l'adresse {ip_address} a changé.\nNuméro de série précédent : {last_serial_number}\nNuméro de série actuel : {current_serial_number}")
        item_id = self.get_item_id(ip_address)
        self.table.item(item_id, tags=('alert',))
        self.remove_printer_sheet(ip_address)

    def create_delete_button(self, ip_address):
        delete_button = ttk.Button(self.table, text='Supprimer la feuille', command=lambda: self.remove_printer_sheet(ip_address))
        delete_button.configure(width=15)
        self.table.window_create('', column='Delete', window=delete_button)


        return None
        
            
    def finalize_report(self):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Récupère la date et l'heure actuelles sous forme de chaîne de caractères
        with open("report.txt", "a") as f:  # Notez le mode 'a' pour ajouter à la fin du fichier
            f.write(f"\n\n--- RAPPORT D'EXÉCUTION - {timestamp} ---\n")  # Ajoute l'horodatage au titre
            f.write("===================\n\n")

            f.write("SUCCÈS :\n")
            for msg in self.report["success"]:
                f.write(f"- {msg}\n")

            f.write("\nERREURS :\n")
            for msg in self.report["errors"]:
                f.write(f"- {msg}\n")